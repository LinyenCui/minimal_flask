#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
測試生成週報表功能
"""

import os
import sys
from datetime import datetime, timedelta
from flask import Flask
from sqlalchemy import create_engine, text
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 創建一個測試用的 Flask 應用
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:0720@localhost:5432/dispatch_db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# 直接使用SQLAlchemy引擎
engine = create_engine(app.config['SQLALCHEMY_DATABASE_URI'])

# 生成週報表函數
def generate_test_weekly_report(category=None):
    try:
        # 獲取日期範圍
        today = datetime.now().date()
        days_since_sunday = today.weekday() + 1
        last_sunday = today - timedelta(days=days_since_sunday + 7)
        last_saturday = last_sunday + timedelta(days=6)
        
        print(f"上週日期範圍: {last_sunday} 至 {last_saturday}")
        
        # 查詢數據
        query_params = {
            "start_date": last_sunday,
            "end_date": last_saturday
        }
        
        query = """
        SELECT 
            ct.id,
            ct.date, 
            ct.start_point, 
            ct.via_point, 
            ct.end_point, 
            ct.meter_fare, 
            ct.extra_fare,
            COALESCE(ct.meter_fare, 0) + COALESCE(ct.extra_fare, 0) as actual_fare,
            ct.driver_id
        FROM 
            completed_trips ct
        WHERE 
            ct.date >= :start_date AND ct.date <= :end_date
        """
        
        if category:
            query += " AND ct.category = :category"
            query_params["category"] = category
            
        query += " ORDER BY ct.date, ct.id"
        
        with engine.connect() as conn:
            result = conn.execute(text(query), query_params)
            completed_trips = result.fetchall()
        
        if not completed_trips:
            category_text = f"類別「{category}」的" if category else ""
            return f"上周沒有{category_text}已完成的班次。", None
        
        # 打印查詢結果
        print(f"找到 {len(completed_trips)} 條記錄:")
        for trip in completed_trips:
            print(trip)
        
        # 創建DataFrame
        df = pd.DataFrame(completed_trips)
        df.columns = ['ID', '日期', '起點', '途經點', '終點', '錶價', '加成', '實收', '司機編號']
        
        # 處理日期和星期
        df['日期'] = pd.to_datetime(df['日期'])
        weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
        df['星期'] = df['日期'].dt.dayofweek.map(weekday_map)
        df['日期'] = df['日期'].dt.strftime('%Y-%m-%d')
        
        # 重排列順序
        df = df[['ID', '日期', '星期', '起點', '途經點', '終點', '司機編號', '錶價', '加成', '實收']]
        
        # 計算總計
        total_meter_fare = df['錶價'].sum()
        total_extra_fare = df['加成'].sum() if df['加成'].notna().any() else 0
        total_actual_fare = df['實收'].sum()
        
        # 按司機分組統計
        driver_stats = df.groupby('司機編號').agg({
            'ID': 'count',
            '實收': 'sum'
        }).reset_index()
        driver_stats.columns = ['司機編號', '班次數', '金額']
        
        # 創建Excel文件
        report_date = datetime.now().strftime('%Y%m%d')
        category_suffix = f"_{category}" if category else ""
        filename = f"weekly_report{category_suffix}_{report_date}.xlsx"
        
        # 直接使用 openpyxl 創建工作簿
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = '班次詳情'
        
        # 添加標題
        title = f"診所班次請款單 ({last_sunday.strftime('%Y/%m/%d')} - {last_saturday.strftime('%Y/%m/%d')})"
        worksheet.cell(row=1, column=1).value = title
        worksheet.merge_cells('A1:J1')
        
        # 添加表頭
        headers = ['ID', '日期', '星期', '起點', '途經點', '終點', '司機編號', '錶價', '加成', '實收']
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=2, column=col_num).value = header
        
        # 添加數據
        for row_num, row_data in enumerate(df.values, 3):
            for col_num, cell_value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num).value = cell_value
        
        # 添加總計行
        total_row = len(df) + 3
        worksheet.cell(row=total_row, column=9).value = "加總:"
        worksheet.cell(row=total_row, column=10).value = total_actual_fare
        
        # 添加司機統計
        driver_stats_row = total_row + 3  # 在總計行下方留出空行
        
        # 添加司機統計標題
        worksheet.cell(row=driver_stats_row, column=1).value = "司機統計:"
        worksheet.cell(row=driver_stats_row, column=1).font = Font(name='微軟正黑體', size=12, bold=True)
        
        # 添加司機統計表頭
        driver_stats_headers = ['司機編號', '班次數', '金額']
        for col_num, header in enumerate(driver_stats_headers, 1):
            cell = worksheet.cell(row=driver_stats_row + 1, column=col_num)
            cell.value = header
            cell.font = Font(name='微軟正黑體', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # 淺藍色
        
        # 添加司機統計數據
        for row_num, row_data in enumerate(driver_stats.values, driver_stats_row + 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                if col_num == 3:  # 金額列
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0'
        
        # 設置列寬
        column_widths = {
            'A': 10,  # ID
            'B': 15,  # 日期
            'C': 8,   # 星期
            'D': 15,  # 起點
            'E': 20,  # 途經點
            'F': 15,  # 終點
            'G': 12,  # 司機編號
            'H': 12,  # 錶價
            'I': 12,  # 加成
            'J': 12,  # 實收
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # 添加基本樣式
        # 標題樣式
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.font = Font(name='微軟正黑體', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        title_cell.font = Font(name='微軟正黑體', size=16, bold=True, color='FFFFFF')
        
        # 表頭樣式
        header_font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col in range(1, len(headers) + 1):
            cell = worksheet.cell(row=2, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 數據行樣式 - 交替行顏色
        for row in range(3, len(df) + 3):
            for col in range(1, len(headers) + 1):
                cell = worksheet.cell(row=row, column=col)
                # 交替行顏色
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='EDF2F9', end_color='EDF2F9', fill_type='solid')  # 淺藍色
                
                # 數字列右對齊
                if col in [8, 9, 10]:  # 錶價、加成、實收列
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 總計行樣式
        total_label = worksheet.cell(row=total_row, column=9)
        total_label.font = Font(name='微軟正黑體', size=12, bold=True)
        total_label.alignment = Alignment(horizontal='right', vertical='center')
        total_label.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # 中藍色
        
        total_cell = worksheet.cell(row=total_row, column=10)
        total_cell.font = Font(name='微軟正黑體', size=12, bold=True)
        total_cell.alignment = Alignment(horizontal='right', vertical='center')
        total_cell.number_format = '#,##0'
        total_cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # 中藍色
        
        # 添加邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 為主表格添加邊框
        for row in range(1, len(df) + 4):  # 包括標題、表頭、數據和總計行
            for col in range(1, len(headers) + 1):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # 為司機統計表添加邊框
        for row in range(driver_stats_row + 1, driver_stats_row + 2 + len(driver_stats)):
            for col in range(1, 4):
                worksheet.cell(row=row, column=col).border = thin_border
        
        # 保存工作簿
        workbook.save(filename)
        
        category_text = f"類別「{category}」的" if category else ""
        return f"已生成上周 ({last_sunday.strftime('%m/%d')} - {last_saturday.strftime('%m/%d')}) {category_text}報表。", filename
        
    except Exception as e:
        return f"生成報表失敗: {str(e)}", None

# 執行測試
print("開始生成週報表...")
result, filename = generate_test_weekly_report()
print(result)

if filename:
    print(f"報表已生成: {filename}")
    # 檢查文件是否存在
    if os.path.exists(filename):
        print(f"文件大小: {os.path.getsize(filename)} 字節")
    else:
        print("警告: 文件未找到!")
else:
    print("報表生成失敗!") 