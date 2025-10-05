"""
報表生成服務模組 - 負責生成週報表並上傳到 Google Drive
"""
import os
import logging
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from flask import current_app
from sqlalchemy import text
from modules.utils.taiwan_time import get_taiwan_time, get_taiwan_date

# 從模型模組導入數據庫連接
from modules.models.base import db

# 導入Google Drive服務
from modules.services.drive_service import upload_file_to_drive

# 建立日誌記錄器
logger = logging.getLogger(__name__)

def _create_excel_report(worksheet, df, driver_stats, filename, start_date, end_date, report_type="週"):
    """
    創建Excel報表的共用函數
    
    Args:
        worksheet: openpyxl worksheet物件
        df: 班次資料DataFrame
        driver_stats: 司機統計DataFrame
        filename: 輸出檔案名
        start_date: 開始日期
        end_date: 結束日期
        report_type: 報表類型（"週"或"月"）
    
    Returns:
        tuple: (結果消息, 生成的文件名)
    """
    # 添加表頭
    headers = ['ID', '日期', '星期', '起點', '途經點', '終點', '司機編號', '錶價', '加成', '實收', '說明']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=2, column=col_num).value = header
    
    # 添加數據
    for row_num, row_data in enumerate(df.values, 3):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
    
    # 計算總計
    total_actual_fare = df['實收'].sum()
    
    # 添加總計行
    total_row = len(df) + 3
    worksheet.cell(row=total_row, column=9).value = "加總:"
    worksheet.cell(row=total_row, column=10).value = total_actual_fare
    
    # 添加司機統計
    driver_stats_row = total_row + 3  # 在總計行下方留出空行
    
    # 簡化下方區域設計 - 使用合併儲存格
    
    # === 司機統計區域 (A-D欄) ===
    # 司機統計標題
    worksheet.cell(row=driver_stats_row, column=1).value = "司機統計"
    worksheet.merge_cells(f'A{driver_stats_row}:D{driver_stats_row}')
    worksheet.cell(row=driver_stats_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=driver_stats_row, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    worksheet.cell(row=driver_stats_row, column=1).font = Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF')  # 字體小一號
    
    # 簡化的司機統計表頭（金額和班次數互換）
    driver_headers = ['司機編號', '金額', '班次數']
    for col_num, header in enumerate(driver_headers, 1):
        cell = worksheet.cell(row=driver_stats_row + 1, column=col_num)
        cell.value = header
        cell.font = Font(name='微軟正黑體', size=10, bold=True, color='FFFFFF')  # 字體小一號
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    
    # 司機統計數據（簡化版，金額和班次數互換）
    for row_num, (_, row_data) in enumerate(driver_stats.iterrows(), driver_stats_row + 2):
        worksheet.cell(row=row_num, column=1).value = row_data['司機編號']
        worksheet.cell(row=row_num, column=2).value = row_data['金額']      # 第二欄改成金額
        worksheet.cell(row=row_num, column=3).value = row_data['班次數']    # 第三欄改成班次數
        
        # 設置格式
        for col in range(1, 4):
            cell = worksheet.cell(row=row_num, column=col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='微軟正黑體', size=9)  # 字體小一號
            if col in [2, 3]:  # 金額、班次數
                cell.number_format = '#,##0'
            # 交替行顏色
            if (row_num - driver_stats_row) % 2 == 0:
                cell.fill = PatternFill(start_color='EDF2F9', end_color='EDF2F9', fill_type='solid')
    
    # === 簽名區域 (E-K欄) ===
    signature_height = max(4, 2 + len(driver_stats))  # 至少4行高度
    
    # 簽名區域標題
    worksheet.cell(row=driver_stats_row, column=5).value = "領款人簽名"
    worksheet.merge_cells(f'E{driver_stats_row}:K{driver_stats_row}')
    worksheet.cell(row=driver_stats_row, column=5).alignment = Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=driver_stats_row, column=5).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    worksheet.cell(row=driver_stats_row, column=5).font = Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF')  # 字體小一號
    
    # 簽名區域內容（大範圍合并）
    signature_area_start = driver_stats_row + 1
    signature_area_end = driver_stats_row + signature_height - 1
    worksheet.merge_cells(f'E{signature_area_start}:K{signature_area_end}')
    
    signature_cell = worksheet.cell(row=signature_area_start, column=5)
    signature_cell.value = "\n\n簽名：\n\n\n日期：    年    月    日"
    signature_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    signature_cell.font = Font(name='微軟正黑體', size=10)  # 字體小一號
    
    # 添加邊框
    stats_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 為司機統計區域添加邊框 (A-D欄)
    for row in range(driver_stats_row, driver_stats_row + signature_height):
        for col in range(1, 5):  # A-D欄
            worksheet.cell(row=row, column=col).border = stats_border
    
    # 為簽名區域添加邊框 (E-K欄)
    for row in range(driver_stats_row, driver_stats_row + signature_height):
        for col in range(5, 12):  # E-K欄
            worksheet.cell(row=row, column=col).border = stats_border
    
    # 設置列寬 - 優化直式A4紙列印
    column_widths = {
        'A': 6,   # ID - 縮小
        'B': 12,  # 日期 - 縮小
        'C': 4,   # 星期 - 縮小
        'D': 12,  # 起點 - 縮小
        'E': 12,  # 途經點 - 大幅縮小
        'F': 12,  # 終點 - 縮小
        'G': 8,   # 司機編號 - 縮小
        'H': 8,   # 錶價 - 縮小
        'I': 8,   # 加成 - 縮小
        'J': 8,   # 實收 - 縮小
        'K': 18,  # 說明欄 - 適中，支援換行
    }
    # 總寬度約: 6+12+4+12+12+12+8+8+8+8+18 = 108，適合A4直式列印
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # 添加基本樣式
    # 標題樣式
    title_cell = worksheet.cell(row=1, column=1)
    title_cell.font = Font(name='微軟正黑體', size=15, bold=True, color='FFFFFF')  # 字體小一號
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # 表頭樣式
    header_font = Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF')  # 字體小一號
    header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col in range(1, len(headers) + 1):
        cell = worksheet.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 數據行樣式 - 交替行顏色
    for row in range(3, len(df) + 3):
        for col in range(1, len(headers) + 1):
            cell = worksheet.cell(row=row, column=col)
            # 設置字體大小
            cell.font = Font(name='微軟正黑體', size=10)  # 字體小一號
            
            # 交替行顏色
            if row % 2 == 0:
                cell.fill = PatternFill(start_color='EDF2F9', end_color='EDF2F9', fill_type='solid')  # 淺藍色
            
            # 數字列右對齊
            if col in [8, 9, 10]:  # 錶價、加成、實收列
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 所有文字欄位都設置換行，必要時調整行高
            if col in [4, 5, 6, 11]:  # 起點、途經點、終點、說明欄
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # 設置行高自動調整（針對含有換行符號或長文字的欄位）
            if cell.value and (str(cell.value).count('\n') > 0 or len(str(cell.value)) > 20):
                # 計算所需行高：基本行高 + 每個換行符號增加高度 + 長文字折行高度
                base_height = 20
                newline_count = str(cell.value).count('\n')
                text_length = len(str(cell.value))
                # 預估折行數（假設每個欄位平均字數）
                estimated_wraps = max(0, (text_length - 20) // 30) if text_length > 20 else 0
                required_height = base_height + (newline_count * 15) + (estimated_wraps * 15)
                
                current_height = worksheet.row_dimensions[row].height or 15
                worksheet.row_dimensions[row].height = max(current_height, required_height, 25)
    
    # 總計行樣式
    total_label = worksheet.cell(row=total_row, column=9)
    total_label.font = Font(name='微軟正黑體', size=11, bold=True)  # 字體小一號
    total_label.alignment = Alignment(horizontal='right', vertical='center')
    total_label.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # 中藍色
    
    total_cell = worksheet.cell(row=total_row, column=10)
    total_cell.font = Font(name='微軟正黑體', size=11, bold=True)  # 字體小一號
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    total_cell.number_format = '#,##0'
    total_cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # 中藍色
    
    # 添加邊框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 為主表格添加邊框
    for row in range(1, len(df) + 4):  # 包括標題、表頭、數據和總計行
        for col in range(1, len(headers) + 1):
            worksheet.cell(row=row, column=col).border = thin_border
    
    # 保存工作簿
    workbook = worksheet.parent
    workbook.save(filename)
    logger.info(f"報表已成功生成: {filename}")
    
    category_text = ""
    return f"已生成上{report_type} ({start_date.strftime('%m/%d')} - {end_date.strftime('%m/%d')}) {category_text}報表。", filename

def generate_monthly_report(category=None):
    """
    生成上個月的統計圖表報表（不是明細表）
    
    Args:
        category: 選擇性的班次類別過濾（例如"診所"或"東洋"）
        
    Returns:
        tuple: (結果消息, 生成的文件名)
    """
    try:
        # 獲取日期範圍 - 上個月的第一天和最後一天
        today = get_taiwan_date()
        # 計算上個月
        if today.month == 1:
            last_month_start = today.replace(year=today.year-1, month=12, day=1)
        else:
            last_month_start = today.replace(month=today.month-1, day=1)
        
        # 計算上個月最後一天
        if last_month_start.month == 12:
            next_month_first = last_month_start.replace(year=last_month_start.year+1, month=1, day=1)
        else:
            next_month_first = last_month_start.replace(month=last_month_start.month+1, day=1)
        last_month_end = next_month_first - timedelta(days=1)
        
        logger.info(f"生成月報表統計圖表，日期範圍: {last_month_start} 至 {last_month_end}")
        
        # 查詢數據
        query_params = {
            "start_date": last_month_start,
            "end_date": last_month_end
        }
        
        query = """
        SELECT 
            ct.date, 
            ct.start_point, 
            ct.end_point, 
            ct.meter_fare, 
            ct.extra_fare,
            COALESCE(ct.meter_fare, 0) + COALESCE(ct.extra_fare, 0) as actual_fare,
            ct.driver_id
        FROM 
            completed_trips ct
        WHERE 
            ct.date >= :start_date AND ct.date <= :end_date
        """
        
        if category:
            if category != "全部":  # 如果不是"全部"，則按類別過濾
                query += " AND ct.category = :category"
                query_params["category"] = category
            logger.info(f"使用類別過濾: {category}")
            
        query += " ORDER BY ct.date, ct.driver_id"
        
        with db.engine.connect() as conn:
            result = conn.execute(text(query), query_params)
            completed_trips = result.fetchall()
        
        if not completed_trips:
            category_text = f"類別「{category}」的" if category and category != "全部" else ""
            logger.warning(f"上個月沒有{category_text}已完成的班次")
            return f"上個月沒有{category_text}已完成的班次。", None
        
        logger.info(f"找到 {len(completed_trips)} 條班次記錄")
        
        # 創建DataFrame進行統計分析
        df = pd.DataFrame(completed_trips)
        df.columns = ['日期', '起點', '終點', '錶價', '加成', '實收', '司機編號']
        df['日期'] = pd.to_datetime(df['日期'])
        
        # 創建Excel工作簿
        report_date = get_taiwan_date().strftime('%Y%m%d')
        category_suffix = f"_{category}" if category and category != "全部" else ""
        filename = f"monthly_stats{category_suffix}_{report_date}.xlsx"
        
        workbook = openpyxl.Workbook()
        
        # 刪除默認工作表，重新創建
        workbook.remove(workbook.active)
        
        # 創建統計報表
        return _create_monthly_statistics_report(workbook, df, filename, last_month_start, last_month_end, category)
        
    except Exception as e:
        logger.error(f"生成月報表失敗: {str(e)}", exc_info=True)
        return f"生成月報表失敗: {str(e)}", None


def _create_monthly_statistics_report(workbook, df, filename, start_date, end_date, category):
    """
    創建月報表統計圖表
    """
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.series import DataPoint
    
    # === 第一個工作表：總覽摘要 ===
    summary_sheet = workbook.create_sheet("月度總覽")
    
    # 標題
    if category == "東洋":
        title_prefix = "東洋班次月度統計報表"
    elif category == "診所":
        title_prefix = "診所班次月度統計報表"
    else:
        title_prefix = "班次月度統計報表"
    
    title = f"{title_prefix} ({start_date.strftime('%Y/%m/%d')} - {end_date.strftime('%Y/%m/%d')})"
    summary_sheet.cell(row=1, column=1).value = title
    summary_sheet.merge_cells('A1:J1')
    
    # 設置標題樣式
    title_cell = summary_sheet.cell(row=1, column=1)
    title_cell.font = Font(name='微軟正黑體', size=16, bold=True, color='FFFFFF')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # 基本統計數據
    total_trips = len(df)
    total_revenue = df['實收'].sum()
    avg_fare = df['實收'].mean()
    unique_drivers = df['司機編號'].nunique()
    
    # 統計摘要表格
    summary_data = [
        ["統計項目", "數值", ""],
        ["總班次數", total_trips, "趟"],
        ["總車資", total_revenue, "元"],
        ["平均車資", round(avg_fare, 0), "元"],
        ["參與司機數", unique_drivers, "人"],
        ["日均班次", round(total_trips / (end_date - start_date).days, 1), "趟/日"]
    ]
    
    # 填入摘要數據
    for row_idx, row_data in enumerate(summary_data, 3):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = summary_sheet.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            
            if row_idx == 3:  # 表頭
                cell.font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.font = Font(name='微軟正黑體', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx == 2 and isinstance(cell_value, (int, float)):  # 數值列
                    cell.number_format = '#,##0'
    
    # === 第二個工作表：司機統計圖表 ===
    driver_sheet = workbook.create_sheet("司機統計")
    
    # 司機統計數據
    driver_stats = df.groupby('司機編號').agg({
        '實收': ['count', 'sum', 'mean']
    }).round(0)
    driver_stats.columns = ['班次數', '總金額', '平均車資']
    driver_stats = driver_stats.reset_index()
    driver_stats = driver_stats.sort_values('總金額', ascending=False)
    
    # 司機統計表格
    driver_sheet.cell(row=1, column=1).value = "司機績效統計"
    driver_sheet.merge_cells('A1:D1')
    driver_sheet.cell(row=1, column=1).font = Font(name='微軟正黑體', size=14, bold=True, color='FFFFFF')
    driver_sheet.cell(row=1, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    driver_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    # 表頭
    headers = ['司機編號', '班次數', '總金額', '平均車資']
    for col_idx, header in enumerate(headers, 1):
        cell = driver_sheet.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 司機數據
    for row_idx, (_, row_data) in enumerate(driver_stats.iterrows(), 3):
        for col_idx, value in enumerate([row_data['司機編號'], row_data['班次數'], row_data['總金額'], row_data['平均車資']], 1):
            cell = driver_sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_idx in [2, 3, 4]:  # 數值列
                cell.number_format = '#,##0'
    
    # 創建司機營收餅圖
    pie_chart = PieChart()
    pie_chart.title = "司機營收分佈"
    
    # 數據範圍
    data_range = Reference(driver_sheet, min_col=3, min_row=2, max_row=2+len(driver_stats), max_col=3)
    labels_range = Reference(driver_sheet, min_col=1, min_row=3, max_row=2+len(driver_stats))
    
    pie_chart.add_data(data_range, titles_from_data=True)
    pie_chart.set_categories(labels_range)
    
    # 設置圖表樣式
    pie_chart.width = 15
    pie_chart.height = 10
    
    # 添加圖表到工作表
    driver_sheet.add_chart(pie_chart, "F3")
    
    # === 第三個工作表：日期趨勢 ===
    trend_sheet = workbook.create_sheet("車資趨勢")
    
    # 按日期統計
    daily_stats = df.groupby(df['日期'].dt.date).agg({
        '實收': ['count', 'sum']
    }).round(0)
    daily_stats.columns = ['班次數', '總金額']
    daily_stats = daily_stats.reset_index()
    
    # 日期趨勢表格
    trend_sheet.cell(row=1, column=1).value = "每日車資趨勢"
    trend_sheet.merge_cells('A1:C1')
    trend_sheet.cell(row=1, column=1).font = Font(name='微軟正黑體', size=14, bold=True, color='FFFFFF')
    trend_sheet.cell(row=1, column=1).fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    trend_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    # 表頭
    trend_headers = ['日期', '班次數', '總金額']
    for col_idx, header in enumerate(trend_headers, 1):
        cell = trend_sheet.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 日期數據
    for row_idx, (_, row_data) in enumerate(daily_stats.iterrows(), 3):
        for col_idx, value in enumerate([row_data['日期'], row_data['班次數'], row_data['總金額']], 1):
            cell = trend_sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_idx in [2, 3]:  # 數值列
                cell.number_format = '#,##0'
    
    # 添加總計行
    total_row = 2 + len(daily_stats) + 1
    trend_sheet.cell(row=total_row, column=2).value = "總計:"
    trend_sheet.cell(row=total_row, column=2).font = Font(name='微軟正黑體', size=11, bold=True)
    trend_sheet.cell(row=total_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
    trend_sheet.cell(row=total_row, column=2).fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    
    trend_sheet.cell(row=total_row, column=3).value = daily_stats['總金額'].sum()
    trend_sheet.cell(row=total_row, column=3).font = Font(name='微軟正黑體', size=11, bold=True)
    trend_sheet.cell(row=total_row, column=3).alignment = Alignment(horizontal='center', vertical='center')
    trend_sheet.cell(row=total_row, column=3).number_format = '#,##0'
    trend_sheet.cell(row=total_row, column=3).fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
    
    # 創建車資趨勢長條圖
    bar_chart = BarChart()
    bar_chart.title = "每日車資趨勢圖"
    bar_chart.x_axis.title = "日期"
    bar_chart.y_axis.title = "車資金額"
    
    # 數據範圍
    data_range = Reference(trend_sheet, min_col=3, min_row=2, max_row=2+len(daily_stats), max_col=3)
    labels_range = Reference(trend_sheet, min_col=1, min_row=3, max_row=2+len(daily_stats))
    
    bar_chart.add_data(data_range, titles_from_data=True)
    bar_chart.set_categories(labels_range)
    
    # 設置圖表樣式
    bar_chart.width = 20
    bar_chart.height = 12
    
    # 添加圖表到工作表
    trend_sheet.add_chart(bar_chart, "E3")
    
    # 設置欄寬
    # 總覽表：加寬欄位以避免文字切割
    summary_sheet.column_dimensions['A'].width = 15  # 統計項目欄
    summary_sheet.column_dimensions['B'].width = 12  # 數值欄
    summary_sheet.column_dimensions['C'].width = 8   # 單位欄
    
    # 司機統計表
    for col in ['A', 'B', 'C', 'D']:
        driver_sheet.column_dimensions[col].width = 12
    
    # 趨勢表：加寬日期欄
    trend_sheet.column_dimensions['A'].width = 12  # 日期欄
    trend_sheet.column_dimensions['B'].width = 10  # 班次數欄
    trend_sheet.column_dimensions['C'].width = 12  # 總金額欄
    
    # 保存工作簿
    workbook.save(filename)
    logger.info(f"月度統計報表已成功生成: {filename}")
    
    category_text = f"類別「{category}」的" if category and category != "全部" else ""
    return f"已生成上月 ({start_date.strftime('%m/%d')} - {end_date.strftime('%m/%d')}) {category_text}統計報表。", filename


def generate_weekly_report(category=None):
    """
    生成上週的班次報表
    
    Args:
        category: 選擇性的班次類別過濾（例如"診所"或"東洋"）
        
    Returns:
        tuple: (結果消息, 生成的文件名)
    """
    try:
        # 獲取日期範圍
        today = get_taiwan_date()
        days_since_sunday = today.weekday() + 1 if today.weekday() < 6 else 0
        last_sunday = today - timedelta(days=days_since_sunday + 7)
        last_saturday = last_sunday + timedelta(days=6)
        
        logger.info(f"生成週報表，日期範圍: {last_sunday} 至 {last_saturday}")
        
        # 查詢數據
        query_params = {
            "start_date": last_sunday,
            "end_date": last_saturday
        }
        
        query = """
        SELECT 
            ct.id,
            ct.date, 
            ct.start_point, 
            ct.via_point, 
            ct.end_point, 
            ct.meter_fare, 
            ct.extra_fare,
            COALESCE(ct.meter_fare, 0) + COALESCE(ct.extra_fare, 0) as actual_fare,
            ct.driver_id,
            ct.modification_reason,
            ct.passenger_leave_reason
        FROM 
            completed_trips ct
        WHERE 
            ct.date >= :start_date AND ct.date <= :end_date
        """
        
        if category:
            if category != "全部":  # 如果不是"全部"，則按類別過濾
                query += " AND ct.category = :category"
                query_params["category"] = category
            logger.info(f"使用類別過濾: {category}")
            
        query += " ORDER BY ct.date, ct.id"
        
        with db.engine.connect() as conn:
            result = conn.execute(text(query), query_params)
            completed_trips = result.fetchall()
        
        if not completed_trips:
            category_text = f"類別「{category}」的" if category and category != "全部" else ""
            logger.warning(f"上周沒有{category_text}已完成的班次")
            return f"上周沒有{category_text}已完成的班次。", None
        
        logger.info(f"找到 {len(completed_trips)} 條班次記錄")
        
        # 創建DataFrame
        df = pd.DataFrame(completed_trips)
        df.columns = ['ID', '日期', '起點', '途經點', '終點', '錶價', '加成', '實收', '司機編號', '修改原因', '請假原因']
        
        # 處理說明欄位：合併修改原因和請假原因
        def combine_remarks(row):
            modification = row['修改原因'] if pd.notna(row['修改原因']) and row['修改原因'].strip() else ""
            leave = row['請假原因'] if pd.notna(row['請假原因']) and row['請假原因'].strip() else ""
            
            if modification and leave:
                return f"{modification}\n{leave}"
            elif modification:
                return modification
            elif leave:
                return leave
            else:
                return ""
        
        df['說明'] = df.apply(combine_remarks, axis=1)
        
        # 處理日期和星期
        df['日期'] = pd.to_datetime(df['日期'])
        weekday_map = {0: '一', 1: '二', 2: '三', 3: '四', 4: '五', 5: '六', 6: '日'}
        df['星期'] = df['日期'].dt.dayofweek.map(weekday_map)
        df['日期'] = df['日期'].dt.strftime('%Y-%m-%d')
        
        # 重排列順序，添加說明欄
        df = df[['ID', '日期', '星期', '起點', '途經點', '終點', '司機編號', '錶價', '加成', '實收', '說明']]
        
        # 計算總計
        total_meter_fare = df['錶價'].sum()
        total_extra_fare = df['加成'].sum() if df['加成'].notna().any() else 0
        total_actual_fare = df['實收'].sum()
        
        # 按司機分組統計
        driver_stats = df.groupby('司機編號').agg({
            'ID': 'count',
            '實收': 'sum'
        }).reset_index()
        driver_stats.columns = ['司機編號', '班次數', '金額']
        
        # 創建Excel文件
        report_date = get_taiwan_date().strftime('%Y%m%d')
        category_suffix = f"_{category}" if category and category != "全部" else ""
        filename = f"weekly_report{category_suffix}_{report_date}.xlsx"
        
        # 直接使用 openpyxl 創建工作簿
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = '班次詳情'
        
        # 根據類別動態生成標題
        if category == "東洋":
            title_prefix = "東洋班次請款單"
        elif category == "診所":
            title_prefix = "診所班次請款單"
        else:
            title_prefix = "班次請款單"
        
        title = f"{title_prefix} ({last_sunday.strftime('%Y/%m/%d')} - {last_saturday.strftime('%Y/%m/%d')})"
        worksheet.cell(row=1, column=1).value = title
        worksheet.merge_cells('A1:K1')  # 擴展到K列，因為多了說明欄
        
        # 使用共用的Excel格式化函數
        return _create_excel_report(worksheet, df, driver_stats, filename, last_sunday, last_saturday, "週")
        
    except Exception as e:
        logger.error(f"生成報表失敗: {str(e)}", exc_info=True)
        return f"生成報表失敗: {str(e)}", None

def upload_to_google_drive(file_path, folder_id=None):
    """
    將文件上傳到Google Drive並設置分享權限
    
    Args:
        file_path: 要上傳的文件路徑
        folder_id: Google Drive文件夾ID（可選）
        
    Returns:
        str: 文件分享鏈接或錯誤消息
    """
    try:
        logger.info(f"嘗試上傳文件到Google Drive: {file_path}")
        
        # 使用drive_service模組中的上傳函數
        success, result = upload_file_to_drive(file_path, folder_id=folder_id)
        
        if success:
            logger.info(f"文件上傳成功，分享鏈接: {result}")
            return result
        else:
            logger.error(f"上傳到Google Drive失敗: {result}")
            return f"上傳到Google Drive失敗: {result}"
    
    except Exception as e:
        logger.error(f"上傳到Google Drive過程中出錯: {str(e)}", exc_info=True)
        return f"上傳到Google Drive失敗: {str(e)}"

def handle_generate_weekly_report(text):
    """
    處理生成周報表命令
    
    Args:
        text: 用戶輸入的文本命令
        
    Returns:
        str: 處理結果消息
    """
    # 解析命令參數
    parts = text.strip().split()
    category = None
    
    if len(parts) > 1:
        category = parts[1]
    
    # 類別與文件夾 ID 的映射
    CATEGORY_FOLDER_MAPPING = {
        "診所": "1Wwp1xIxnn9m9qlvX_BwpE30K0AgLVdYe",  # 診所文件夾 ID
        "東洋": "1dctU8QPRWNPn57LxpcYTeKKcsGn_dLOU",   # 東洋文件夾 ID
        "全部": None  # 全部類別不指定特定文件夾
    }
    
    # 根據類別獲取對應的文件夾ID
    folder_id = None
    if category and category in CATEGORY_FOLDER_MAPPING:
        folder_id = CATEGORY_FOLDER_MAPPING[category]
        logger.info(f"使用類別: {category}, 對應文件夾ID: {folder_id}")
    
    try:
        logger.info(f"開始生成週報表，類別: {category}")
        # 生成報表
        result, filename = generate_weekly_report(category)
        
        if not filename:
            logger.warning("沒有生成報表文件")
            return result
        
        # 上傳到Google Drive
        logger.info(f"報表生成成功，準備上傳到Google Drive: {filename}")
        drive_url = upload_to_google_drive(filename, folder_id)
        
        if drive_url and not drive_url.startswith("上傳到Google Drive失敗"):
            logger.info(f"報表已成功上傳: {drive_url}")
            return f"{result}\n報表已上傳到Google Drive: {drive_url}"
        else:
            logger.warning(f"報表上傳失敗: {drive_url}")
            return f"{result}\n報表已生成，但上傳到Google Drive失敗: {drive_url}"
    except Exception as e:
        logger.error(f"處理生成週報表命令時出錯: {str(e)}", exc_info=True)
        return f"生成報表時出錯: {str(e)}"


def handle_generate_monthly_report(text):
    """
    處理生成月報表命令
    
    Args:
        text: 用戶輸入的文本命令 (格式: /生成月報表 診所 [YYYY-MM])
        
    Returns:
        str: 處理結果消息
    """
    # 解析命令參數
    parts = text.strip().split()
    category = None
    target_month = None
    
    if len(parts) > 1:
        category = parts[1]
    
    # 檢查是否有指定年月參數
    if len(parts) > 2:
        try:
            from datetime import datetime
            month_str = parts[2]
            target_month = datetime.strptime(month_str, '%Y-%m').date()
            logger.info(f"指定目標月份: {target_month}")
        except ValueError:
            logger.warning(f"無效的月份格式: {parts[2]}, 使用預設月份")
    
    # 如果沒有指定月份，使用上一個完整月份
    if not target_month:
        from modules.utils.taiwan_time import get_taiwan_date
        today = get_taiwan_date()
        if today.month == 1:
            target_month = today.replace(year=today.year-1, month=12, day=1)
        else:
            target_month = today.replace(month=today.month-1, day=1)
        logger.info(f"使用預設月份: {target_month}")
    
    # 類別與文件夾 ID 的映射（與週報表相同）
    CATEGORY_FOLDER_MAPPING = {
        "診所": "1Wwp1xIxnn9m9qlvX_BwpE30K0AgLVdYe",  # 診所文件夾 ID
        "東洋": "1dctU8QPRWNPn57LxpcYTeKKcsGn_dLOU",   # 東洋文件夾 ID
        "全部": None  # 全部類別不指定特定文件夾
    }
    
    # 根據類別獲取對應的文件夾ID
    folder_id = None
    if category and category in CATEGORY_FOLDER_MAPPING:
        folder_id = CATEGORY_FOLDER_MAPPING[category]
        logger.info(f"使用類別: {category}, 對應文件夾ID: {folder_id}")
    
    try:
        logger.info(f"開始生成月報表，類別: {category}, 月份: {target_month}")
        
        # 如果是診所類別，嘗試使用新的月結單功能
        if category == "診所":
            try:
                from modules.services.google_sheets_service import create_monthly_statement_sheet
                
                # 創建月結單Google Sheets
                sheets_url = create_monthly_statement_sheet(target_month, category)
                
                if sheets_url:
                    logger.info(f"月結單試算表創建成功: {sheets_url}")
                    return f"✅ 月結單已生成完成！\n\n📊 試算表連結: {sheets_url}\n\n💡 包含月結單封面、司機統計、車資趨勢等詳細資訊"
                else:
                    logger.warning("月結單試算表創建失敗，降級到Excel報表")
                    raise Exception("Google Sheets創建失敗")
                    
            except Exception as e:
                logger.warning(f"月結單功能不可用，降級到Excel報表: {str(e)}")
                # 降級到原有的Excel報表功能
                result, filename = generate_monthly_report(category)
                
                if not filename:
                    logger.warning("沒有生成報表文件")
                    return result
                
                # 上傳到Google Drive
                logger.info(f"月報表生成成功，準備上傳到Google Drive: {filename}")
                drive_url = upload_to_google_drive(filename, folder_id)
                
                if drive_url and not drive_url.startswith("上傳到Google Drive失敗"):
                    logger.info(f"月報表已成功上傳: {drive_url}")
                    return f"{result}\n報表已上傳到Google Drive: {drive_url}\n\n⚠️ 注意：Google Sheets API未啟用，已降級到Excel報表"
                else:
                    logger.warning(f"月報表上傳失敗: {drive_url}")
                    return f"{result}\n報表已生成，但上傳到Google Drive失敗: {drive_url}\n\n⚠️ 注意：Google Sheets API未啟用，已降級到Excel報表"
        else:
            # 其他類別使用原有的Excel報表功能
            result, filename = generate_monthly_report(category)
            
            if not filename:
                logger.warning("沒有生成報表文件")
                return result
            
            # 上傳到Google Drive
            logger.info(f"月報表生成成功，準備上傳到Google Drive: {filename}")
            drive_url = upload_to_google_drive(filename, folder_id)
            
            if drive_url and not drive_url.startswith("上傳到Google Drive失敗"):
                logger.info(f"月報表已成功上傳: {drive_url}")
                return f"{result}\n報表已上傳到Google Drive: {drive_url}"
            else:
                logger.warning(f"月報表上傳失敗: {drive_url}")
                return f"{result}\n報表已生成，但上傳到Google Drive失敗: {drive_url}"
                
    except Exception as e:
        logger.error(f"處理生成月報表命令時出錯: {str(e)}", exc_info=True)
        return f"生成月報表時出錯: {str(e)}" 