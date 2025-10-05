"""
月結單封面生成器
負責生成Excel格式的月結單封面
"""

import logging
from datetime import date, datetime
from typing import Dict, Any, List, Tuple, Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from modules.models.base import db
from modules.utils.taiwan_time import get_taiwan_date, get_taiwan_time

logger = logging.getLogger(__name__)

def create_monthly_statement_cover(workbook: Workbook, start_date: date, end_date: date, category: str) -> Dict[str, Any]:
    """
    創建月結單封面
    
    Args:
        workbook: Excel工作簿
        start_date: 開始日期
        end_date: 結束日期
        category: 班次類別
        
    Returns:
        Dict: 包含總金額等統計數據
    """
    try:
        # 獲取或創建月結單封面工作表
        if "月結單（封面）" in workbook.sheetnames:
            cover_sheet = workbook["月結單（封面）"]
            cover_sheet.delete_rows(1, cover_sheet.max_row)
        else:
            cover_sheet = workbook.create_sheet("月結單（封面）")
        
        # 將封面工作表移到第一位（openpyxl 3.x 版本的方法）
        try:
            workbook.move(cover_sheet, 0)
        except AttributeError:
            # 如果move方法不存在，使用其他方法
            # 先刪除現有工作表，再創建新的
            if "月結單（封面）" in workbook.sheetnames:
                workbook.remove(workbook["月結單（封面）"])
            cover_sheet = workbook.create_sheet("月結單（封面）", 0)
        
        # 獲取統計數據
        stats = _get_monthly_statistics(start_date, end_date, category)
        
        # 生成結單號碼
        statement_no = f"STMT-{start_date.strftime('%Y%m')}"
        
        # 設定頁面格式
        _setup_page_formatting(cover_sheet)
        
        # 設定列寬
        _set_column_widths(cover_sheet)
        
        # 寫入標題
        _write_header(cover_sheet, statement_no)
        
        # 寫入左側資訊
        _write_left_info(cover_sheet, start_date, end_date, stats)
        
        # 寫入右側金額
        _write_right_amounts(cover_sheet, stats, start_date)
        
        # 寫入可選區塊
        _write_optional_sections(cover_sheet, stats)
        
        logger.info(f"月結單封面生成完成，總金額: {stats['total_amount']:,}")
        return stats
        
    except Exception as e:
        logger.error(f"創建月結單封面失敗: {str(e)}", exc_info=True)
        raise

def render_statement_cover(ws, meta: Dict[str, Any]):
    """
    專責畫封面的函式
    
    Args:
        ws: Excel工作表
        meta: 包含所有必要資訊的字典
    """
    try:
        # 設定頁面格式
        _setup_page_formatting(ws)
        
        # 設定列寬
        _set_column_widths(ws)
        
        # 寫入標題
        _write_header(ws, meta.get('statement_no', 'STMT-000000'))
        
        # 寫入左側資訊
        _write_left_info_from_meta(ws, meta)
        
        # 寫入右側金額
        _write_right_amounts_from_meta(ws, meta)
        
        # 寫入可選區塊
        _write_optional_sections_from_meta(ws, meta)
        
        logger.info(f"月結單封面渲染完成，總金額: {meta.get('total_amount', 0):,}")
        
    except Exception as e:
        logger.error(f"渲染月結單封面失敗: {str(e)}", exc_info=True)
        raise

def _get_monthly_statistics(start_date: date, end_date: date, category: str) -> Dict[str, Any]:
    """獲取月度統計數據（使用與現有月報表完全一致的邏輯）"""
    try:
        from sqlalchemy import text
        
        # 使用與現有月報表完全一致的查詢邏輯
        query_params = {
            "start_date": start_date,
            "end_date": end_date
        }
        
        query = """
        SELECT 
            ct.date, 
            ct.start_point, 
            ct.end_point, 
            ct.meter_fare, 
            ct.extra_fare,
            COALESCE(ct.meter_fare, 0) + COALESCE(ct.extra_fare, 0) as actual_fare,
            ct.driver_id
        FROM 
            completed_trips ct
        WHERE 
            ct.date >= :start_date AND ct.date <= :end_date
        """
        
        if category and category != "全部":
            query += " AND ct.category = :category"
            query_params["category"] = category
        
        query += " ORDER BY ct.date, ct.driver_id"
        
        with db.engine.connect() as conn:
            result = conn.execute(text(query), query_params)
            trips = result.fetchall()
        
        if not trips:
            return {
                'total_amount': 0,
                'total_trips': 0,
                'drivers_top3': [],
                'deposits': [],
                'bank_name': '',
                'last4_mask': '＊＊＊＊',
                'payee_name': '—'
            }
        
        # 計算總金額（與現有月報表完全一致的邏輯）
        total_amount = sum(row[5] for row in trips if row[5] is not None)
        
        # 司機統計
        driver_stats = {}
        for row in trips:
            driver_id = row[6]
            amount = row[5] or 0
            if driver_id in driver_stats:
                driver_stats[driver_id] += amount
            else:
                driver_stats[driver_id] = amount
        
        # 前3名司機
        drivers_sorted = sorted(driver_stats.items(), key=lambda x: x[1], reverse=True)
        drivers_top3 = [(f"司機{driver_id}", int(amount)) for driver_id, amount in drivers_sorted[:3]]
        
        # 獲取入金記錄和銀行資訊
        deposits, bank_name, last4_mask = _get_deposits_and_bank_info(start_date, end_date)
        
        return {
            'total_amount': int(total_amount),
            'total_trips': len(trips),
            'drivers_top3': drivers_top3,
            'deposits': deposits,
            'bank_name': bank_name,
            'last4_mask': last4_mask,
            'payee_name': '—'  # 可從系統設定讀取
        }
        
    except Exception as e:
        logger.error(f"獲取月度統計失敗: {str(e)}")
        return {
            'total_amount': 0,
            'total_trips': 0,
            'drivers_top3': [],
            'deposits': [],
            'bank_name': '',
            'last4_mask': '＊＊＊＊',
            'payee_name': '—'
        }

def _get_deposits_and_bank_info(start_date: date, end_date: date) -> Tuple[List[Dict[str, Any]], str, str]:
    """獲取入金記錄和銀行資訊"""
    try:
        from sqlalchemy import text
        
        query = """
        SELECT occurred_at, amount_in, bank_name, bank_account_last4
        FROM account_ledger 
        WHERE type = 'deposit' 
          AND occurred_at >= :start_time 
          AND occurred_at <= :end_time
        ORDER BY occurred_at DESC 
        LIMIT 5
        """
        
        # 轉換為台北時區的開始和結束時間
        start_datetime = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=ZoneInfo("Asia/Taipei"))
        end_datetime = datetime.combine(end_date, datetime.max.time()).replace(tzinfo=ZoneInfo("Asia/Taipei"))
        
        with db.engine.connect() as conn:
            result = conn.execute(text(query), {
                'start_time': start_datetime,
                'end_time': end_datetime
            })
            deposits = result.fetchall()
        
        deposits_list = []
        bank_name = ''
        last4_mask = '＊＊＊＊'
        
        if deposits:
            # 取最常見的銀行名稱和後四碼
            bank_counts = {}
            last4_counts = {}
            
            for dep in deposits:
                deposits_list.append({
                    'date': dep[0].strftime('%Y/%m/%d') if dep[0] else '',
                    'amount': int(dep[1]) if dep[1] else 0,
                    'bank_name': dep[2] or '',
                    'last4': dep[3] or ''
                })
                
                if dep[2]:
                    bank_counts[dep[2]] = bank_counts.get(dep[2], 0) + 1
                if dep[3]:
                    last4_counts[dep[3]] = last4_counts.get(dep[3], 0) + 1
            
            # 取最常見的銀行名稱
            if bank_counts:
                bank_name = max(bank_counts.items(), key=lambda x: x[1])[0]
            
            # 取最常見的後四碼
            if last4_counts:
                last4_mask = max(last4_counts.items(), key=lambda x: x[1])[0]
        
        return deposits_list, bank_name, last4_mask
        
    except Exception as e:
        logger.error(f"獲取入金記錄和銀行資訊失敗: {str(e)}")
        return [], '', '＊＊＊＊'

def _setup_page_formatting(sheet):
    """設定頁面格式（A4直向，邊界0.8cm）"""
    try:
        # 設定頁面邊界（A4、直向，0.8cm = 0.3英寸）
        sheet.page_margins = PageMargins(
            left=0.3, right=0.3, top=0.3, bottom=0.3,
            header=0.3, footer=0.3
        )
        
        # 設定列印範圍
        sheet.print_area = 'A1:F30'
        
        # 關閉網格線
        sheet.sheet_view.showGridLines = False
        
        # 設定縮放為一頁寬
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        
        # 設定A4直向
        sheet.page_setup.paperSize = 9  # A4
        sheet.page_setup.orientation = 'portrait'
        
    except Exception as e:
        logger.warning(f"設定頁面格式失敗: {str(e)}")

def _set_column_widths(sheet):
    """設定列寬（單位：chars）"""
    try:
        column_widths = {
            'A': 26, 'B': 18, 'C': 10, 
            'D': 22, 'E': 14, 'F': 22
        }
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
            
    except Exception as e:
        logger.warning(f"設定列寬失敗: {str(e)}")

def _write_header(sheet, statement_no: str):
    """寫入標題"""
    try:
        # 主標題
        sheet['A1'] = '派班系統 月結單'
        sheet.merge_cells('A1:C1')
        
        title_cell = sheet['A1']
        title_cell.font = Font(name='Calibri', size=20, bold=True, color='000000')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # 結單號碼
        sheet['F1'] = f'結單號碼：{statement_no}'
        sheet['F1'].font = Font(name='Calibri', size=10, color='666666')
        sheet['F1'].alignment = Alignment(horizontal='right', vertical='center')
        
    except Exception as e:
        logger.error(f"寫入標題失敗: {str(e)}")

def _write_left_info(sheet, start_date: date, end_date: date, stats: Dict[str, Any]):
    """寫入左側資訊"""
    try:
        # 付款人
        sheet['A3'] = '付款人'
        sheet['B3'] = '達恩診所'
        
        # 受款人
        sheet['A4'] = '受款人'
        sheet['B4'] = stats.get('payee_name', '—')
        
        # 帳戶
        sheet['A5'] = '帳戶'
        bank_name = stats.get('bank_name', '')
        last4 = stats.get('last4_mask', '＊＊＊＊')
        if bank_name:
            sheet['B5'] = f'{bank_name} ＊＊＊＊{last4}'
        else:
            sheet['B5'] = f'＊＊＊＊{last4}'
        
        # 期間
        sheet['A6'] = '期間'
        period_text = f'{start_date.year}年{start_date.month}月1日 – {end_date.year}年{end_date.month}月{end_date.day}日'
        sheet['B6'] = period_text
        
        # 備註
        sheet['A8'] = '備註'
        sheet['B8'] = '金額將自上列受款帳戶扣除'
        
        # 格式化左側標籤
        for row in range(3, 9):
            if row != 7:  # 跳過空行
                label_cell = sheet[f'A{row}']
                value_cell = sheet[f'B{row}']
                
                label_cell.font = Font(name='Calibri', size=11, color='666666')
                value_cell.font = Font(name='Calibri', size=12)
                
                if row == 8:  # 備註
                    value_cell.font = Font(name='Calibri', size=11, color='666666')
        
    except Exception as e:
        logger.error(f"寫入左側資訊失敗: {str(e)}")

def _write_left_info_from_meta(sheet, meta: Dict[str, Any]):
    """從meta寫入左側資訊"""
    try:
        # 付款人
        sheet['A3'] = '付款人'
        sheet['B3'] = '達恩診所'
        
        # 受款人
        sheet['A4'] = '受款人'
        sheet['B4'] = meta.get('payee_name', '—')
        
        # 帳戶
        sheet['A5'] = '帳戶'
        bank_name = meta.get('bank_name', '')
        last4 = meta.get('last4_mask', '＊＊＊＊')
        if bank_name:
            sheet['B5'] = f'{bank_name} ＊＊＊＊{last4}'
        else:
            sheet['B5'] = f'＊＊＊＊{last4}'
        
        # 期間
        sheet['A6'] = '期間'
        month_start = meta.get('month_start')
        month_end = meta.get('month_end')
        if month_start and month_end:
            period_text = f'{month_start.year}年{month_start.month}月1日 – {month_end.year}年{month_end.month}月{month_end.day}日'
            sheet['B6'] = period_text
        
        # 備註
        sheet['A8'] = '備註'
        sheet['B8'] = '金額將自上列受款帳戶扣除'
        
        # 格式化左側標籤
        for row in range(3, 9):
            if row != 7:  # 跳過空行
                label_cell = sheet[f'A{row}']
                value_cell = sheet[f'B{row}']
                
                label_cell.font = Font(name='Calibri', size=11, color='666666')
                value_cell.font = Font(name='Calibri', size=12)
                
                if row == 8:  # 備註
                    value_cell.font = Font(name='Calibri', size=11, color='666666')
        
    except Exception as e:
        logger.error(f"從meta寫入左側資訊失敗: {str(e)}")

def _write_right_amounts(sheet, stats: Dict[str, Any], start_date: date):
    """寫入右側金額"""
    try:
        # 總金額標籤
        sheet['D3'] = '總金額 (TWD)'
        sheet['D3'].font = Font(name='Calibri', size=11, color='666666')
        
        # 總金額（大數字）
        total_text = f'NT$ {stats["total_amount"]:,}'
        sheet['D4'] = total_text
        sheet.merge_cells('D4:F7')
        
        amount_cell = sheet['D4']
        amount_cell.font = Font(name='Calibri', size=28, bold=True, color='000000')
        amount_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 分隔線（淡灰底線）
        for col in ['D', 'E', 'F']:
            cell = sheet[f'{col}8']
            cell.border = Border(bottom=Side(style='thin', color='E0E0E0'))
        
        # 月份
        sheet['D9'] = '月份'
        sheet['E9'] = f'{start_date.year}年{start_date.month}月'
        sheet.merge_cells('E9:F9')
        
        # 列印日期
        sheet['D10'] = '列印日期'
        today = get_taiwan_date()
        sheet['E10'] = today.strftime('%Y/%m/%d')
        sheet.merge_cells('E10:F10')
        
        # 格式化
        for row in range(9, 11):
            sheet[f'D{row}'].font = Font(name='Calibri', size=11, color='666666')
            sheet[f'E{row}'].font = Font(name='Calibri', size=12)
            sheet[f'E{row}'].alignment = Alignment(horizontal='left')
        
    except Exception as e:
        logger.error(f"寫入右側金額失敗: {str(e)}")

def _write_right_amounts_from_meta(sheet, meta: Dict[str, Any]):
    """從meta寫入右側金額"""
    try:
        # 總金額標籤
        sheet['D3'] = '總金額 (TWD)'
        sheet['D3'].font = Font(name='Calibri', size=11, color='666666')
        
        # 總金額（大數字）
        total_amount = meta.get('total_amount', 0)
        total_text = f'NT$ {total_amount:,}'
        sheet['D4'] = total_text
        sheet.merge_cells('D4:F7')
        
        amount_cell = sheet['D4']
        amount_cell.font = Font(name='Calibri', size=28, bold=True, color='000000')
        amount_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 分隔線（淡灰底線）
        for col in ['D', 'E', 'F']:
            cell = sheet[f'{col}8']
            cell.border = Border(bottom=Side(style='thin', color='E0E0E0'))
        
        # 月份
        sheet['D9'] = '月份'
        month_start = meta.get('month_start')
        if month_start:
            sheet['E9'] = f'{month_start.year}年{month_start.month}月'
            sheet.merge_cells('E9:F9')
        
        # 列印日期
        sheet['D10'] = '列印日期'
        printed_on = meta.get('printed_on')
        if printed_on:
            sheet['E10'] = printed_on.strftime('%Y/%m/%d')
        else:
            today = get_taiwan_date()
            sheet['E10'] = today.strftime('%Y/%m/%d')
        sheet.merge_cells('E10:F10')
        
        # 格式化
        for row in range(9, 11):
            sheet[f'D{row}'].font = Font(name='Calibri', size=11, color='666666')
            sheet[f'E{row}'].font = Font(name='Calibri', size=12)
            sheet[f'E{row}'].alignment = Alignment(horizontal='left')
        
    except Exception as e:
        logger.error(f"從meta寫入右側金額失敗: {str(e)}")

def _write_optional_sections(sheet, stats: Dict[str, Any]):
    """寫入可選區塊"""
    try:
        # 入金摘要（左下）
        if stats['deposits']:
            sheet['A14'] = '入金摘要'
            sheet['A14'].font = Font(name='Calibri', size=11, bold=True, color='666666')
            
            for i, deposit in enumerate(stats['deposits'][:3], 15):
                deposit_text = f'{deposit["date"]}  NT${deposit["amount"]:,}（{deposit.get("last4", "N/A")}）'
                sheet[f'A{i}'] = deposit_text
                sheet[f'A{i}'].font = Font(name='Calibri', size=11)
        
        # 司機摘要（右下）
        if stats['drivers_top3']:
            sheet['D12'] = '司機摘要'
            sheet['D12'].font = Font(name='Calibri', size=11, bold=True, color='666666')
            
            for i, (driver_name, amount) in enumerate(stats['drivers_top3'], 13):
                driver_text = f'{driver_name}  NT${amount:,}'
                sheet[f'D{i}'] = driver_text
                sheet[f'D{i}'].font = Font(name='Calibri', size=11)
        
    except Exception as e:
        logger.error(f"寫入可選區塊失敗: {str(e)}")

def _write_optional_sections_from_meta(sheet, meta: Dict[str, Any]):
    """從meta寫入可選區塊"""
    try:
        # 入金摘要（左下）
        deposits = meta.get('deposits', [])
        if deposits:
            sheet['A14'] = '入金摘要'
            sheet['A14'].font = Font(name='Calibri', size=11, bold=True, color='666666')
            
            for i, deposit in enumerate(deposits[:3], 15):
                deposit_text = f'{deposit["date"]}  NT${deposit["amount"]:,}（{deposit.get("last4", "N/A")}）'
                sheet[f'A{i}'] = deposit_text
                sheet[f'A{i}'].font = Font(name='Calibri', size=11)
        
        # 司機摘要（右下）
        drivers_top3 = meta.get('drivers_top3', [])
        if drivers_top3:
            sheet['D12'] = '司機摘要'
            sheet['D12'].font = Font(name='Calibri', size=11, bold=True, color='666666')
            
            for i, (driver_name, amount) in enumerate(drivers_top3, 13):
                driver_text = f'{driver_name}  NT${amount:,}'
                sheet[f'D{i}'] = driver_text
                sheet[f'D{i}'].font = Font(name='Calibri', size=11)
        
    except Exception as e:
        logger.error(f"從meta寫入可選區塊失敗: {str(e)}")